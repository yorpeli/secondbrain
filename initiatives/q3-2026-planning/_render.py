#!/usr/bin/env python3
# Q3 2026 planning renderer (canonical). Reads the two workbooks, auto-maps line-items to the 9 threads,
# applies discussion decisions, emits plan.md + q3-plan-review.html. Re-runnable. Lives in the initiative folder.
import openpyxl, re, html

BASE = "/Users/yorpeli/Documents/Dev/SecondBrain/initiatives/q3-2026-planning"
CLM  = f"{BASE}/docs/CLM Quarter Planning Q3 2026.xlsx"
SS   = f"{BASE}/docs/H2 2026 Candidates - Self Service.xlsx"

THREADS = [
  (1,"Regulatory & licensing (AI moat)","Meital (Licenses, exec cover for Yael) · Sitara (PM) · Ido (P&Q/India) · License Infra","deadline","India/PACB P0 + moat base-layer"),
  (2,"Vendor modernization & orchestration","Elad (EVS/DU)","protected","KYC approval rate / efficiency (+ QA layers)"),
  (3,"Approval CVR & Activation (→FFT)","KYC Journey/FDC + Self-Service","growth","Approval CVR (T1/T2) + FFT rate / time-to-FFT"),
  (4,"Existing-base migration / CLM cleanup","Ido (+ Estella AHA delta)","deadline","100% compliance by Jan'27; 98% rev retention"),
  (5,"Partners / Enterprise","Estella + Eliya","growth","Enterprise/partner CVR (eBay renewal)"),
  (6,"AI KYC / Agentic","Yonatan / Shilhav","strategic","60% auto-resolution on D-leads (Jul 15); flagship"),
  (7,"China TICP optimization","Jojo Zhou + CLM China","growth","China TICP CVR / lead-rank accuracy"),
  (8,"Localization","L&L (confirm)","growth","Localized onboarding CVR per country"),
  (9,"Modernization / post-rollout platform","Engineering: Yonatan + Avital (on-prem→cloud, .NET, data arch; Monitoring & Alerting here)","modernization","~40% · engineering-led; feeds AI/business outcomes"),
  (10,"Top-of-funnel: Mobile & Lead Scoring","Self-Service (Ira) · dep: Mobile team (external)","growth","Mobile signup→submission CVR; TICP lead-rank"),
]
TYPE_LABEL={"deadline":"deadline","protected":"protected growth","growth":"growth","strategic":"strategic","modernization":"~40% · modernization"}
TYPE_CLASS={"deadline":"t-deadline","protected":"t-protected","growth":"t-growth","strategic":"t-strategic","modernization":"t-mod"}

SHEET_DEFAULT={"Licenses":1,"P&Q Plan":1,"EVS Q3 plan":2,"KYC Journey Q3 2026":3,"FDC Q3 2026":3,"DU Q3 2026":2,"License Infra":1,"India License":8,"H2 2026 Candidates - Self Servi":3}
NAME_KEYS=["item name","initiative","key result","item"]
COLMAP={"priority":["priority"],"mw":["man weeks effort","men weeks effort","mw effort","est mw"],"committed":["committed"],"metric":["top line metric"],"cat":["category","area"],"desc":["item description"],"owner":["owner"],"deps":["dependencies in other domains","dependencies internal/external","dependencies on other domains","dependencies"]}
STOP={"draft mode","draft","total sum of mw","select...","select…","na",""}
SKIP_RE=re.compile(r"^(total sum|comitted|committed sum|capacity|over capacity|available manpower|mws in q|tech debt|buffer deduction|extra holiday|draft mode|sprint \d|qa$|<item|<task|dev\d|total effort)",re.I)
# pure-modernization -> primary thread 9
MODPRIM_RE=re.compile(r"state machine|orchestrator|refactor|aurora|config(uration)? cleanup|engineering excellence|management tools|deprecat|sign document|data collection manager|ocr extraction implementation|unified kyc dashboard|\bdashboard\b|new rfi center|new fdc configuration",re.I)
# modernization-flavoured (cross-tag 9 if primary elsewhere) -> counts toward the 40%
MODX_RE=re.compile(r"to cloud|data.?driven routing|modernization|legacy|automation coverage|monitoring",re.I)

def norm(s): return "" if s is None else str(s).replace("\n"," / ").strip()
def prio(s):
  m=re.search(r"P[0-4]",str(s or "")); return m.group(0) if m else ""

def extract(path):
  wb=openpyxl.load_workbook(path,data_only=True); rows=[]
  for ws in wb.worksheets:
    if ws.title not in SHEET_DEFAULT: continue
    header=None;hi=0
    for i,r in enumerate(ws.iter_rows(values_only=True),1):
      cells=[norm(c).lower() for c in r]
      if any(any(k==c or k in c for k in NAME_KEYS) for c in cells if c): header=cells;hi=i;break
    if not header: continue
    def col(keys):
      for k in keys:
        for ci,h in enumerate(header):
          if h==k or (h and k in h): return ci
      return None
    ni=None
    for k in NAME_KEYS:
      for ci,h in enumerate(header):
        if h==k: ni=ci;break
      if ni is not None: break
    if ni is None: continue
    cols={f:col(ks) for f,ks in COLMAP.items()}
    for r in list(ws.iter_rows(values_only=True))[hi:]:
      def g(f):
        ci=cols.get(f); v=norm(r[ci]) if (ci is not None and ci<len(r)) else ""
        return "" if v.lower().startswith("select") else v
      name=norm(r[ni]) if ni<len(r) else ""; low=name.lower()
      if low in STOP or len(name)<3 or low.startswith("select") or low.startswith("<") or "total sum" in low or SKIP_RE.match(low): continue
      rows.append({"tab":ws.title,"name":name,"priority":prio(g("priority")) or prio(name),"mw":g("mw"),"committed":g("committed"),"metric":g("metric"),"cat":g("cat"),"desc":g("desc"),"owner":g("owner"),"deps":g("deps")})
  return rows

def classify(it):
  t=(it["name"]+" "+it["desc"]+" "+it["cat"]+" "+it["tab"]).lower(); cross=set(); deps=[]
  prim=SHEET_DEFAULT[it["tab"]]
  if re.search(r"\bchina\b|cn lead|china model",t): prim=7
  elif re.search(r"lead score|lead rank|lead scoring|\bmobile\b|signup|sign[- ]?up|top of funnel|prepopulate phone|block old devices",t): prim=10
  elif re.search(r"localization|localize|localized|singpass|singapore|\buae\b|quebec|\bkr\b|turkey|ukraine|google address",t): prim=8
  elif re.search(r"partner|enterprise|consent|ebay|lite plan|\bwl\b|4step",t): prim=5
  elif re.search(r"agentic|ai agent|agent that does",t): prim=6
  elif re.search(r"\baha\b|clm cleanup|clean ?up|existing (customer|population)|periodic review|\bcrp\b|payer.*receiver|migration of existing",t): prim=4
  elif MODPRIM_RE.search(t): prim=9
  # cross tags
  if prim!=9 and (MODPRIM_RE.search(t) or MODX_RE.search(t)): cross.add(9)
  if prim!=4 and re.search(r"migrat|migration",t): cross.add(4)
  if re.search(r"add company|person entity|person level|entity model|new entity|legal entity collection|person via vendor",t): cross.add("A")
  if prim!=2 and re.search(r"\bvendor\b|au10tix|aiprise|sumsub|trulioo|persona|smart.*routing|selfie vendor|resistant",t): cross.add(2)
  if prim!=8 and re.search(r"\bindia\b",t): cross.add(8)
  if re.search(r"\bps team\b|order ps|open ps|payoneer service",t): deps.append("PS")
  if it["deps"]: deps.append(it["deps"][:90])
  cross.discard(prim)
  return prim,[str(c) for c in sorted(cross,key=str)],deps

rows=extract(CLM)+extract(SS)
for it in rows:
  p,c,d=classify(it); it["thread"]=p; it["cross"]=c; it["depslist"]=d
# manual additions from discussion
rows.append({"tab":"discussion","name":"eBay Renewal + KYC Optimization","priority":"P0","mw":"","committed":"","metric":"Protect+grow eBay partner rev (renewal)","cat":"Enterprise","desc":"","owner":"Estella/Eliya","deps":"","thread":5,"cross":[],"depslist":[]})
rows.append({"tab":"Delegated Onboarding","name":"Onboarding Handover","priority":"","mw":"","committed":"No (UNFUNDED)","metric":"","cat":"","desc":"at-risk; scope TBD","owner":"Meital","deps":"","thread":5,"cross":["1"],"depslist":[]})
rows.append({"tab":"Enterprise deck","name":"Unified KYC Dashboard","priority":"P1","mw":"","committed":"Yes","metric":"Internal KYC visibility","cat":"Modernization","desc":"On clm-main Looker semantic layer","owner":"analytics/data","deps":"","thread":9,"cross":[],"depslist":["clm-main semantic layer"]})

by={t[0]:[] for t in THREADS}
for it in rows: by[it["thread"]].append(it)

POSTURE=[
 "**Strategic frame: Licensing as an AI Moat** (Yaron + Yonatan, top-down). A compliant licensed base + AI-driven KYX automation compounds into a defensible advantage. **Top-line outcome = FFT** (approval CVR -> activation -> FFT). **T1/T2 company** leverage lens across all threads.",
 "**Protect a growth floor**, weighted to **vendor modernization** (freed localization dev -> EVS/DU).",
 "**Compromise localization** -> Thread 8, deprioritized (India committed; SG/UAE at-risk; US finishing with local vendors).",
 "**Self-Service protects BRR** (Business Ready Rate), not CVR. \"Open PS during onboarding\" -> P1.",
 "**Deadline threads (1 & 4) take first call** on capacity.",
 "**Modernization (Thread 9) ~= 40% of effort**, cross-cutting. It overlaps the protected vendor-mod floor (vendors-to-cloud counts in both), so the 40% is *not* fully additive on top of the growth floor.",
]
INBOUND=[
 ("Cards","Multi-entity work","Cross-cut A (entity/person model): DEM person-level, Entity Model [Person], person eKYX, Add company","🟢 aligned (consolidate our 4 scattered items)"),
 ("Product Compliance","Multi-entity work","Cross-cut A (same)","🟢 aligned"),
 ("Money-Movement-Platform","VOP (Verification of Payee)","License Infra VOP (P2)","🟠 in plan at P2 - reconcile priority"),
]
DECK=[
 ("01","KYC Services Monitoring + one-time existing-pop cleanup (~$10M/yr; eBay/CN)","Thread 4 + License Infra customer-tagging Monitoring + eBay→T5","✅ cleanup covered; external validation of Thread 4 ($ + Compliance sponsor)"),
 ("02a","Consent flow - push consent after document submission","Thread 5 consent page + SS fast-follower","✅ already have it; reinforces the P0 (~50% drop)"),
 ("02b","Billing-country role-based locking (companies)","SS 'HQ vs incorporation mapping' (P1)","~partial"),
 ("03","Expose Final CRP for Walmart","'Migrate CRP Prep Work' = CRP deprecation","❌ net-new + TENSION (we're deprecating CRP) — PENDING"),
 ("04","AHA audit risky-payee attempts (~$40.5M TFL, 15.2K AHs)","adjacent to AHA-CLM-delta, different (risk/compliance)","❌ net-new — PENDING"),
 ("05","Payees Journey Control (gate activation + comms transfer to Enterprise)","not in plan","❌ DECLINED — no surface-ownership moves either direction; no branched flows for onboarded customers yet"),
 ("06","Unified KYC Dashboard","clm-main Looker semantic layer (built)","✅ COMMITTED — internal teams dashboard, low effort → Thread 9"),
]
DECISIONS=[
 "9 threads locked. Modernization Thread 9 (~40%) is PRIMARILY ENGINEERING-LED infra (on-prem->cloud, .NET/framework upgrades, data-component re-architecture); product items are a subset (⊕9). Separate full modernization deck coming from engineering.",
 "Multi-entity consumers = Cards + Product Compliance (corrected from Fin-Crimes). eBay = a RENEWAL (partner contract), P0.",
 "Exec deck for Oren built (RUN/GROW · CLM HL · Q3 Initiatives · Risks · Dependencies) - captured in the DB memory doc.",
 "Reconciled with Yaron's top-down CLM Q3 Investment Plan + Yonatan's response (the strategic spine). Frame: Licensing as an AI Moat; FFT top-line; T1/T2 lens. Docs in docs/.",
 "Added Thread 10 (Top-of-funnel: Mobile & Lead Scoring) - front half of funnel; external Mobile-team dependency. AI KYC raised to 60% auto-resolution. Modernization co-owned Yonatan + Avital; Monitoring under Modernization; Cost-to-serve distributed (T6 + T2). CSP unfunded + rollout-stabilization = risks; InLife out of scope.",
 "Licenses owner = Meital (exec cover for Yael, on leave); Sitara (Principal PM) executes. (Correction 2026-06-20: an earlier 'Yonatan Birger leads Licenses' note was an error - Birger is a Senior PM on Elad's KYC team.)",
 "Partners owners = Estella + Eliya; eBay KYC Optimization added P0.",
 "Posture: protect growth, weighted to vendor modernization.",
 "Self-Service protects BRR; \"Open PS during onboarding\" P2->P1.",
 "Delegated Onboarding: \"Onboarding Handover\" added, at-risk/unfunded.",
 "Yael on leave; Meital exec cover - temp.",
 "Enterprise ask 05 (Payees Journey Control) DECLINED - no surface-ownership transfers between CLM and Enterprise; no branched flows for onboarded customers yet.",
 "Enterprise ask 06 (Unified KYC Dashboard) COMMITTED - Thread 9, on clm-main semantic layer.",
 "Cards + Fin-Crimes request multi-entity (cross-cut A); MMP requests VOP (consider bump from P2).",
]

# ---------------- plan.md ----------------
def md_items(items):
  if not items: return "_no items mapped yet_\n"
  o=["| P | Item | Source | MW | Committed | Top metric | Cross | Deps |","|---|------|--------|----|-----------|-----------|-------|------|"]
  for it in sorted(items,key=lambda x:(x["priority"] or "P9",x["name"])):
    o.append(f"| {it['priority'] or '-'} | {it['name'][:70]} | {it['tab']} | {it['mw'] or '-'} | {it['committed'] or '-'} | {it['metric'] or '-'} | {' '.join('⊕'+c for c in it['cross']) or '-'} | {', '.join(it['depslist']) or '-'} |")
  return "\n".join(o)+"\n"
M=[f"# Q3 2026 CLM Plan - working draft\n",
   "> **Living working doc** (canonical: re-render with `python3 _render.py`). `q3-plan-review.html` is generated from the same script. DB memory doc refreshed at checkpoints.",
   "> Window: 2026-07-01 to 2026-09-30. Owner: Yonatan. Agent: q-plan-pm.",
   f"> First-pass auto-map ({len(rows)} items). ⊕ = cross-thread tag (A = entity/person model; digits = other threads, incl ⊕9 = counts toward the ~40% modernization). REVIEW + correct misfits.\n",
   "## Build steps",
   "1. [x] Lock the main threads (9 threads + cross-cut A)",
   "2. [~] **Map planning items -> threads (first-pass auto-map; needs review)**",
   "3. [~] Cross-team dependencies + inbound requests (below)",
   "4. [ ] Map everything to top-line metrics (workbook metric carried per item)\n",
   "## Posture (decided)"]+[f"- {p}" for p in POSTURE]+["\n## Threads overview","| # | Thread | Owner | Type | Items |","|---|--------|-------|------|-------|"]
for tid,name,owner,typ,metric in THREADS: M.append(f"| {tid} | {name} | {owner} | {TYPE_LABEL[typ]} | {len(by[tid])} |")
M.append("\n- **Cross-cut A:** entity / person-level \"add company\" model (Cards + Product Compliance both pulling on it)\n")
for tid,name,owner,typ,metric in THREADS:
  M.append(f"\n## Thread {tid} - {name}\n**Owner:** {owner} · **Type:** {TYPE_LABEL[typ]} · **Top-line:** {metric}\n")
  if tid==8: M.append("> Country status: **India** (start, committed) · **Singapore + UAE** (at-risk) · **US** (finishing, local vendors).\n")
  if tid==9: M.append("> Cross-cutting: ⊕9-tagged items elsewhere also count toward the ~40% modernization target (e.g. vendors-to-cloud in Thread 2).\n")
  M.append(md_items(by[tid]))
M.append("\n## Cross-team requests (inbound demand)\n")
M.append("| Requesting team | Ask | Maps to | Status |\n|---|---|---|---|")
for a,b,c,d in INBOUND: M.append(f"| **{a}** | {b} | {c} | {d} |")
M.append("\n### Platform-Enterprise deck (Gal Appel) - mapping only\n")
M.append("| # | Enterprise ask | Maps to | Verdict |\n|---|---|---|---|")
for a,b,c,d in DECK: M.append(f"| {a} | {b} | {c} | {d} |")
M.append("\n## Open gaps, risks & data asks")
M+=["- **RISK - CSP / Delegated Onboarding** (the \"Onboarding Handover\"): a CVR lever per Yaron (Tier-1 hubs, CSP→customer handover) but currently **UNFUNDED**. Track as a risk.",
    "- **RISK - Rollout stabilization:** post-rollout defect remediation is the principal Q3 capacity constraint; a big enough rollout issue would eat all other efforts. Sequence against it, don't layer on top.",
    "- **OUT OF SCOPE - Ongoing lifecycle / InLife** (re-verification, reactivation, expansion beyond add-company): explicitly not in Q3 - called out as a known omission.",
    "- **Dependency - Mobile team** (external to CLM) for the Top-of-funnel / mobile track (Thread 10).",
    "- Covered, not gaps: Cost-to-serve = AI-KYC manual-review reduction (T6) + vendor-orchestration costs (T2); Monitoring & Alerting = under Modernization (T9).",
    "- No MW estimates for Licenses & KYC Journey; DU capacity = `#REF!`; License Infra over capacity (21/19).",
    "- Existing-base migration gap analysis (which segments fail which CLM requirements) net-new, unscoped.",
    "- Enterprise asks 03 (Walmart CRP) + 04 (AHA audit) PENDING decision.\n"]
M.append("## Decisions log")
M+=[f"[2026-06-20] {d}" for d in DECISIONS]
open(f"{BASE}/plan.md","w").write("\n".join(M)+"\n")

# ---------------- HTML ----------------
def esc(s): return html.escape(str(s or ""))
def rows_html(items):
  if not items: return '<tr><td colspan="8" class="tbd">no items mapped yet</td></tr>'
  r=""
  for it in sorted(items,key=lambda x:(x["priority"] or "P9",x["name"])):
    pc={"P0":"p0","P1":"p1","P2":"p2","P3":"p3","P4":"p4"}.get(it["priority"],"")
    cross=" ".join(f'<span class="xt">⊕{esc(c)}</span>' for c in it["cross"])
    r+=f'<tr><td><span class="pri {pc}">{esc(it["priority"] or "-")}</span></td><td class="nm">{esc(it["name"])}</td><td class="src">{esc(it["tab"])}</td><td>{esc(it["mw"] or "-")}</td><td>{esc(it["committed"] or "-")}</td><td>{esc(it["metric"] or "-")}</td><td>{cross or "-"}</td><td class="dep">{", ".join(esc(d) for d in it["depslist"]) or "-"}</td></tr>'
  return r
tabs='<button class="tab active" onclick="show(this,0)">Overview</button>'
tabs+="".join(f'<button class="tab" onclick="show(this,{t[0]})">{t[0]}. {esc(t[1].split(" ")[0])}… <span class="cnt">{len(by[t[0]])}</span></button>' for t in THREADS)
tabs+='<button class="tab" onclick="show(this,99)">Cross-team</button>'
ov="".join(f'<tr><td class="num">{t[0]}</td><td class="tname">{esc(t[1])}</td><td>{esc(t[2])}</td><td><span class="tag {TYPE_CLASS[t[3]]}">{TYPE_LABEL[t[3]]}</span></td><td>{esc(t[4])}</td><td class="cnt2">{len(by[t[0]])}</td></tr>' for t in THREADS)
panels=f'<div class="panel active" id="p0"><h3>Overview · {len(rows)} items</h3><table class="ov"><thead><tr><th class="num">#</th><th>Thread</th><th>Owner</th><th>Type</th><th>Top-line metric</th><th>Items</th></tr></thead><tbody>{ov}</tbody></table><p class="note">⊕ = cross-thread tag · A = entity/person model · ⊕9 = counts toward the ~40% modernization. First-pass auto-map — review.</p><div class="post"><b>Posture:</b><ul>'+"".join(f"<li>{esc(re.sub(chr(92)+'*','',p))}</li>" for p in POSTURE)+'</ul></div></div>'
for tid,name,owner,typ,metric in THREADS:
  note=""
  if tid==8: note='<p class="note">Country status: <b>India</b> (committed) · <b>Singapore + UAE</b> (at-risk) · <b>US</b> (finishing, local vendors).</p>'
  if tid==9: note='<p class="note">Cross-cutting: ⊕9-tagged items in other threads also count toward the ~40% modernization target.</p>'
  panels+=f'<div class="panel" id="p{tid}"><h3>{tid}. {esc(name)} <span class="tag {TYPE_CLASS[typ]}">{TYPE_LABEL[typ]}</span></h3><div class="ph">Owner: <b>{esc(owner)}</b> · Top-line: <b>{esc(metric)}</b></div>{note}<table><thead><tr><th>P</th><th>Item</th><th>Source</th><th>MW</th><th>Committed</th><th>Top metric</th><th>Cross</th><th>Deps</th></tr></thead><tbody>{rows_html(by[tid])}</tbody></table></div>'
inb="".join(f'<tr><td class="tname">{esc(a)}</td><td>{esc(b)}</td><td>{esc(c)}</td><td>{esc(d)}</td></tr>' for a,b,c,d in INBOUND)
dk="".join(f'<tr><td class="num">{esc(a)}</td><td>{esc(b)}</td><td>{esc(c)}</td><td>{esc(d)}</td></tr>' for a,b,c,d in DECK)
panels+=f'<div class="panel" id="p99"><h3>Cross-team demand</h3><h4>Inbound requests</h4><table><thead><tr><th>Team</th><th>Ask</th><th>Maps to</th><th>Status</th></tr></thead><tbody>{inb}</tbody></table><h4 style="margin-top:18px">Platform-Enterprise deck (Gal Appel) — mapping only</h4><table><thead><tr><th>#</th><th>Ask</th><th>Maps to</th><th>Verdict</th></tr></thead><tbody>{dk}</tbody></table></div>'

SHELL=r'''<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Q3 2026 CLM Plan - Review</title>
<style>
:root{--bg:#0f1115;--panel:#171a21;--panel2:#1e222b;--line:#2a2f3a;--ink:#e8eaed;--muted:#9aa3b2;--accent:#6ea8fe;--deadline:#ff6b6b;--growth:#51cf66;--protected:#20c997;--strategic:#b197fc;--mod:#4dabf7;--warn:#ffd43b}
*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--ink);font:14px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Arial,sans-serif}
.wrap{max-width:1180px;margin:0 auto;padding:26px 22px 60px}
h1{margin:0 0 4px;font-size:24px;letter-spacing:-.3px}.sub{color:var(--muted);font-size:13px}h4{margin:0 0 8px;font-size:13px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em}
.meta{margin:10px 0 4px;display:flex;gap:16px;flex-wrap:wrap;color:var(--muted);font-size:12px}.meta b{color:var(--ink)}
.tabs{display:flex;gap:6px;flex-wrap:wrap;margin:18px 0 16px;border-bottom:1px solid var(--line);padding-bottom:10px}
.tab{background:var(--panel2);border:1px solid var(--line);color:var(--muted);border-radius:8px;padding:6px 11px;font-size:12.5px;cursor:pointer}
.tab:hover{color:var(--ink)}.tab.active{color:var(--ink);border-color:var(--accent);box-shadow:0 0 0 1px var(--accent) inset}
.cnt{background:var(--bg);border-radius:10px;padding:0 6px;margin-left:3px;font-size:11px}
.panel{display:none}.panel.active{display:block}
h3{font-size:17px;margin:6px 0 8px}.ph{color:var(--muted);font-size:13px;margin-bottom:10px}
table{width:100%;border-collapse:collapse;background:var(--panel);border:1px solid var(--line);border-radius:10px;overflow:hidden;margin-bottom:6px}
th,td{padding:8px 10px;text-align:left;vertical-align:top;border-bottom:1px solid var(--line);font-size:13px}
th{background:var(--panel2);color:var(--muted);font-size:10.5px;text-transform:uppercase;letter-spacing:.06em}
tr:last-child td{border-bottom:none}.num{color:var(--muted);width:30px}.tname{font-weight:600}.nm{font-weight:500}
.src{color:var(--muted);font-size:12px}.dep{color:var(--muted);font-size:12px}.cnt2{font-variant-numeric:tabular-nums;color:var(--accent);font-weight:600}
.note{color:var(--muted);font-size:12px;margin:8px 0 0}.tbd{color:var(--muted)}
.post{background:var(--panel);border:1px solid var(--line);border-radius:10px;padding:10px 16px;margin-top:14px}.post ul{margin:6px 0 0;padding-left:18px}.post li{margin:3px 0;font-size:13px}
.tag{display:inline-block;padding:2px 8px;border-radius:6px;font-size:11.5px;font-weight:600}
.t-deadline{background:rgba(255,107,107,.13);color:var(--deadline)}.t-growth{background:rgba(81,207,102,.13);color:var(--growth)}
.t-protected{background:rgba(32,201,151,.15);color:var(--protected)}.t-strategic{background:rgba(177,151,252,.15);color:var(--strategic)}.t-mod{background:rgba(77,171,247,.15);color:var(--mod)}
.pri{display:inline-block;min-width:24px;text-align:center;padding:1px 5px;border-radius:5px;font-size:11.5px;font-weight:700;color:#0f1115}
.p0{background:#ff6b6b}.p1{background:#ffa94d}.p2{background:#ffd43b}.p3{background:#74c0fc}.p4{background:#495057;color:#cbd3df}
.xt{display:inline-block;background:var(--panel2);border:1px solid var(--line);color:var(--mod);border-radius:5px;padding:0 5px;font-size:11px;margin-right:2px}
</style></head><body><div class="wrap">
<h1>Q3 2026 CLM Plan</h1><div class="sub">Strategic threads + auto-mapped items - review draft (first post-rollout quarter)</div>
<div class="meta"><span>Window: <b>Jul 1 - Sep 30, 2026</b></span><span>Owner: <b>Yonatan Orpeli</b></span><span>Threads: <b>9</b> + cross-cut A</span><span>Items: <b>__N__</b></span></div>
<div class="tabs">__TABS__</div>
__PANELS__
</div>
<script>function show(btn,i){document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));document.getElementById('p'+i).classList.add('active');btn.classList.add('active');}</script>
</body></html>'''
out=SHELL.replace("__N__",str(len(rows))).replace("__TABS__",tabs).replace("__PANELS__",panels)
open(f"{BASE}/q3-plan-review.html","w").write(out)
print(f"items: {len(rows)}")
for t in THREADS: print(f"  {t[0]} {t[1][:34]:34} {len(by[t[0]])}")
